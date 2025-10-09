import math
import os
import re
from tkinter import messagebox
import logging as log
from typing import Any
from openpyxl import load_workbook, Workbook
import converter as cnv

INPUT_SHEET = "Input"
REPORT_SHEET = "Report"
DILATION_SHEET = "Dilation"
EIEUR_SHEET = "Ei-Eur"
RFKN_SHEET = "Rf-K-n"
FLAC_SHEET = "FLAC1"
OUTPUT_SHEET = "Summary by Type"
SORT_COLUMN_IDX = cnv.col_name_to_idx("CG")
LAST_COLUMN_IDX = SORT_COLUMN_IDX

MAPPINGS = [
    # sheet name, src cell 1, [src cell 2,] dest col, row offset[, condition]
    {"sheet": INPUT_SHEET, "in1": "D6", "out": "B", "offset": 0},
    {"sheet": INPUT_SHEET, "in1": "D8", "out": "D", "offset": 0},
    {"sheet": INPUT_SHEET, "in1": "I10", "out": "F", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "L6", "out": "G", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "F8", "out": "H", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "L8", "out": "I", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "M8", "out": "J", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "N8", "out": "K", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "F21", "out": "L", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "G21", "out": "L", "offset": 1},
    {"sheet": REPORT_SHEET, "in1": "H21", "out": "L", "offset": 2},
    {"sheet": REPORT_SHEET, "in1": "F20", "out": "T", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "G20", "out": "T", "offset": 1},
    {"sheet": REPORT_SHEET, "in1": "H20", "out": "T", "offset": 2},
    {"sheet": REPORT_SHEET, "in1": "F23", "out": "U", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "G23", "out": "U", "offset": 1},
    {"sheet": REPORT_SHEET, "in1": "H23", "out": "U", "offset": 2},
    {"sheet": REPORT_SHEET, "in1": "L10", "out": "AA", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "M10", "out": "AB", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "N10", "out": "AC", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "O10", "out": "AD", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "L11", "out": "AE", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "Q171", "out": "AG", "offset": 0, "if": "L11", "is": "CU"},
    {"sheet": REPORT_SHEET, "in1": "Q170", "out": "AH", "offset": 0, "if": "L11", "is": "CU"},
    {"sheet": REPORT_SHEET, "in1": "Q164", "in2": "Q195", "out": "AI", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "Q163", "in2": "Q194", "out": "AJ", "offset": 0},
    {"sheet": INPUT_SHEET, "in1": "D11", "out": "AN", "offset": 0},
    {"sheet": DILATION_SHEET, "in1": "V4", "out": "AQ", "offset": 0},
    {"sheet": DILATION_SHEET, "in1": "V30", "out": "AQ", "offset": 1},
    {"sheet": DILATION_SHEET, "in1": "V51", "out": "AQ", "offset": 2},
    {"sheet": DILATION_SHEET, "in1": "V11", "out": "AR", "offset": 0},
    {"sheet": DILATION_SHEET, "in1": "V37", "out": "AR", "offset": 1},
    {"sheet": DILATION_SHEET, "in1": "V58", "out": "AR", "offset": 2},
    {"sheet": REPORT_SHEET, "in1": "F45", "out": "AS", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "G45", "out": "AS", "offset": 1},
    {"sheet": REPORT_SHEET, "in1": "H45", "out": "AS", "offset": 2},
    {"sheet": REPORT_SHEET, "in1": "F46", "out": "AZ", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "G46", "out": "AZ", "offset": 1},
    {"sheet": REPORT_SHEET, "in1": "H46", "out": "AZ", "offset": 2},
    {"sheet": REPORT_SHEET, "in1": "F39", "out": "BB", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "G39", "out": "BB", "offset": 1},
    {"sheet": REPORT_SHEET, "in1": "H39", "out": "BB", "offset": 2},
    {"sheet": EIEUR_SHEET, "in1": "Y2", "out": "BC", "offset": 0},
    {"sheet": EIEUR_SHEET, "in1": "Y24", "out": "BC", "offset": 1},
    {"sheet": EIEUR_SHEET, "in1": "Y49", "out": "BC", "offset": 2},
    {"sheet": EIEUR_SHEET, "in1": "Y19", "out": "BE", "offset": 0},
    {"sheet": EIEUR_SHEET, "in1": "Y42", "out": "BE", "offset": 1},
    {"sheet": EIEUR_SHEET, "in1": "Y67", "out": "BE", "offset": 2},
    {"sheet": EIEUR_SHEET, "in1": "Y9", "out": "BH", "offset": 0},
    {"sheet": EIEUR_SHEET, "in1": "Y31", "out": "BH", "offset": 1},
    {"sheet": EIEUR_SHEET, "in1": "Y57", "out": "BH", "offset": 2},
    {"sheet": RFKN_SHEET, "in1": "G56", "out": "BS", "offset": 0},
    {"sheet": RFKN_SHEET, "in1": "G57", "out": "BT", "offset": 0},
    {"sheet": RFKN_SHEET, "in1": "G75", "out": "BU", "offset": 0},
    {"sheet": RFKN_SHEET, "in1": "G76", "out": "BV", "offset": 0},
    {"sheet": FLAC_SHEET, "in1": "AD22", "out": "BW", "offset": 0},
    {"sheet": FLAC_SHEET, "in1": "AD23", "out": "BW", "offset": 1},
    {"sheet": FLAC_SHEET, "in1": "AD24", "out": "BW", "offset": 2},
    {"sheet": INPUT_SHEET, "in1": "D11", "out": "CG", "offset": 0},
    {"sheet": INPUT_SHEET, "in1": "D11", "out": "CG", "offset": 1},
    {"sheet": INPUT_SHEET, "in1": "D11", "out": "CG", "offset": 2},
]

MERGE_COLS = [
    "B", "D", "F", "G", "H", "I", "J", "K", "AA", "AB", "AC", "AD", "AE", "AG", "AH", "AI", "AJ", "AN", "BS", "BT", "BU", "BV"
]

class ExcelWorkbook:
    """
    This is a wrapper for 'openpyxl' workbook
    """
    def __init__(self, pyxl: Workbook):
        self.pyxl = pyxl


    def __getattr__(self, name):
        return getattr(self.pyxl, name)


def extract_file(source: str, destination: str):
    """Extracts data from one excel file into another in according to mapping rules."""
    log.info("Extracting data from file `%s` to `%s` ...", source, destination)
    wb_out = open_workbook(destination)

    _extract_one(source, wb_out)

    save_workbook(wb_out, destination)
    log.info("Done file extracting")


def extract_dir(source: str, destination: str):
    """Extracts data from all excel files in source directory to destination file in according to mapping rules."""
    log.info("Extracting data from directory '%s' to '%s' ...", source, destination)
    wb_out = open_workbook(destination)
    start_row = max_row(wb_out) + 1

    # Get list of *.xlsx files
    excel_files = _list_excel_files(source)
    if len(excel_files) == 0:
        log.warning("No Excel files found in '%s'", source)
        return

    for fn in excel_files:
        _extract_one(fn, wb_out)

    end_row = max_row(wb_out)
    sort_rows(wb_out, start_row, end_row)

    save_workbook(wb_out, destination)
    log.info("Done directory extracting")


def open_workbook(filename: str) -> ExcelWorkbook:
    """Opens output Excel file."""
    wb = load_workbook(filename)
    wb.active = wb[OUTPUT_SHEET]
    return ExcelWorkbook(wb)


def max_row(wb: ExcelWorkbook) -> int:
    """Returns index of the last workbook row"""
    return wb.pyxl.active.max_row


def insert_one_value(val: Any, wb: ExcelWorkbook, row: int, column: str):
    """Inserts a value into Excel workbook in specified cell

        Args:
            val - value
            wb_out - Excel wrapper
            row - row at which to insert
            column - column identifier, like 'B'
    """
    wb.pyxl.active.cell(
        row = row,
        column = cnv.col_name_to_idx(column),
        value = cnv.convert_na(val)
    )


def save_workbook(wb: ExcelWorkbook, destination: str):
    """Saves workbook into a file specified as a path."""
    if not _rename_orig(destination):
        return

    try:
        wb.pyxl.save(destination)
    except Exception as e: # pylint: disable=broad-except
        log.error("Failed to save file '%s': %s", destination, e)
        os.rename(_new_file_name(destination), destination)


# Based on https://stackoverflow.com/questions/44767554/sorting-with-openpyxl
# Copying format: https://stackoverflow.com/questions/45433425/how-to-move-cell-range-in-openpyxl-with-its-properties-hyperlinks-formatting-e
def sort_rows(wb: ExcelWorkbook, row_start: int, row_end: int):
    """ Sorts range in the sheet

        Args:
            row_start   First row to be sorted
            row_end     Last row to be sorted (inclusive)
    """
    log.info("Sorting rows %s - %s in the output spreadsheet", row_start, row_end)

    ws = wb.pyxl.active

    col_name_start = cnv.col_idx_to_name(1)
    col_name_end = cnv.col_idx_to_name(LAST_COLUMN_IDX)

    org_range = col_name_start + str(row_start) + ':' + col_name_end + str(row_end)
    shift = ws.max_row + 1 - row_start

    # Move whole range to the bottom
    ws.move_range(org_range, rows = shift)

    keys=[]
    for i in range(row_start + shift, row_end + shift + 1):
        val = ws.cell(row=i, column=SORT_COLUMN_IDX).value or ""
        keys.append((val, i))

    # Sorts by all fields in the tuple!
    keys.sort()

    # Move rows from original place to destination in sorted order
    to_row = row_start
    for key in keys:
        row_range = col_name_start + str(key[1]) + ':' + col_name_end + str(key[1])
        ws.move_range(row_range, rows = to_row - key[1])
        to_row += 1


def merge_cells(wb_out: ExcelWorkbook, col: str, last_row: int):
    """Merges three rows in Excel spreadsheet from row 'last_row" exclusively."""
    col_idx = cnv.col_name_to_idx(col)
    wb_out.pyxl.active.merge_cells(start_row=last_row + 1, end_row=last_row + 3, start_column=col_idx, end_column=col_idx)


# Tries twice
def _rename_orig(destination):
    try:
        os.rename(destination, _new_file_name(destination))
        return True
    except PermissionError:
        resp = messagebox.askretrycancel(message=f"File '{destination}' is opened in another application.\n" \
                                    "Either close other and retry or cancel")
        if resp == messagebox.CANCEL:
            return False
        try:
            os.rename(destination, _new_file_name(destination))
            return True
        except PermissionError:
            return False


def _list_excel_files(dir_path: str):
    excel_files = []
    for fn in os.listdir(dir_path):
        full_path = os.path.join(dir_path, fn)
        if re.match(r'.+\.xls[bmx]?$', fn) and os.path.isfile(full_path):
            excel_files.append(full_path)

    return excel_files


def _extract_one(source: str, wb_out: ExcelWorkbook):
    wb_in = load_workbook(source, data_only=True)

    last_row = max_row(wb_out)
    log.info("Extracting data from '%s' from row %d ...", source, last_row)

    for mapping in MAPPINGS:
        _copy_one_value(wb_in, wb_out.pyxl, last_row, mapping)

    for col in MERGE_COLS:
        merge_cells(wb_out, col, last_row)


def _new_file_name(fname: str):
    base, ext = os.path.splitext(fname)
    return base + ".org" + ext


def _print_mappings():
    for mapping in MAPPINGS:
        entry = ''
        if "if" in mapping:
            entry += "IF '" + mapping["if"] + "' == " + mapping["is"] + ": "
        entry += mapping["sheet"] + " (" + mapping["in1"]
        if "in2" in mapping:
            entry += "+" + mapping["in2"] + ")/2 -> " + mapping["out"]
        else:
            entry += ") -> " + mapping["out"]
        print(entry)


def _get_cell_value(wb, sheet, cell):
    try:
        v = wb[sheet][cell].value
        if v is None:
            log.warning("No value in sheet '%s', cell '%s'", sheet, cell)
    except KeyError as e:
        log.warning("Failed to extract value from sheet '%s', cell '%s': %s", sheet, cell, e)
        v = math.nan
    return v


def _is_number(val):
    kind = type(val)
    return kind == int or kind == float and not math.isnan(val)


def _copy_one_value(wb_in: Workbook, wb_out: Workbook, last_row: int, mapping: Any):
    v = cnv.may_be_convert(_get_cell_value(wb_in, mapping["sheet"], mapping["in1"]))

    if "in2" in mapping:
        v2 = cnv.may_be_convert(_get_cell_value(wb_in, mapping["sheet"], mapping["in2"]))
        if _is_number(v) and _is_number(v2):
            v = (v + v2) / 2.0

    v_conv = cnv.convert_na(v)

    if "if" not in mapping or v_conv == cnv.N_A or _get_cell_value(wb_in, mapping["sheet"], mapping["if"]) == mapping["is"]:
        wb_out.active.cell(
            row = last_row + mapping["offset"] + 1,
            column = cnv.col_name_to_idx(mapping["out"]),
            value = v_conv
        )
