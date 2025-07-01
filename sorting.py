from openpyxl import Worksheet


def sort_cells(ws: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int, sort_col: int):
    
    #1 copy cells
    range_cells = []
    for i in range(start_row, end_row):
        row_cells = []
        for j in range(start_col, end_col):
            row_cells.append(ws.cell(row=i, column=j))
    
    #2
    range_rows = []
    for row in ws.iter_rows(min_row=2, max_row=end_row, min_col=start_col, max_col=end_col):
        range_rows.append(row)

    # sort cells
    range_cells.sort(key=lambda item: item[sort_col-start_col].value)


# From https://stackoverflow.com/questions/44767554/sorting-with-openpyxl
def sheet_sort_rows(ws: Worksheet, row_start: int, row_end: int, cols=None, sorter=None):
    """ Sorts given rows of the sheet
        row_start   First row to be sorted
        row_end     Last row to be sorted (default last row)
        cols        Columns to be considered in sort
        sorter      Function that accepts a tuple of values and
                    returns a sortable key
    """

    bottom = ws.max_row
    right = get_column_letter(ws.max_column)
    if cols is None:
        cols = range(1, ws.max_column+1)

    array = {}
    for row in range(row_start, row_end+1):
        key = []
        for col in cols:
            key.append(ws.cell(row, col).value)
        array[key] = array.get(key, set()).union({row})

    order = sorted(array, key=sorter)

    ws.move_range(f"A{row_start}:{right}{row_end}", bottom)
    dest = row_start
    for src_key in order:
        for row in array[src_key]:
            src = row + bottom
            dist = dest - src
            ws.move_range(f"A{src}:{right}{src}", dist) # UC??
            dest += 1