import unittest
import os
# from unittest.mock import Mock, patch
import fitz # PyMuPDF
from openpyxl import load_workbook, Workbook
# import converter as cnv
import pdf_extractor as pdf


class TestExtractor(unittest.TestCase):

    def setUp(self):
        # Create dummy Excel files for testing
        self.source_file = "test_source.pdf"
        self.destination_file = "test_destination.xlsx"


        with fitz.open() as doc:
            page = doc.new_page(width=595, height=842)
            page.insert_text((50, 50), "Client:")
            page.insert_text((110, 50), "Client Name")
            page.insert_text((200, 50), "Borehole/Sample No.:")
            page.insert_text((350, 50), "ENG-PQ24-15 / SC40")
            page.insert_text((200, 70), "Sample Depth (m):")
            page.insert_text((350, 70), "56.9 - 57.6m")
            page.insert_text((200, 90), "Liquid Limit:")
            page.insert_text((350, 90), "39")
            doc.save(self.source_file)

        wb_destination = Workbook()
        ws_destination = wb_destination.active
        ws_destination.title = "Summary by Type"
        wb_destination.save(self.destination_file)


    def tearDown(self):
        # Clean up test files
        # if os.path.exists(self.source_file):
        #     os.remove(self.source_file)
        # if os.path.exists(self.destination_file):
        #     os.remove(self.destination_file)
        pass


    def test_validate_mappings(self):
        for mapping in pdf.MAPPINGS:
            print(mapping)
            self.assertTrue(mapping["key"].islower())
            self.assertTrue(mapping["out"].isalpha())
            self.assertGreaterEqual(mapping["num"], len(mapping["offset"]))


    def test_extract_file(self):
        pdf.extract_file(self.source_file, self.destination_file)

        #UC wb_in = load_workbook(self.source_file, data_only=True)
