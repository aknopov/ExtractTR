import math
import os
import re
from tkinter import messagebox
import logging as log
from openpyxl import load_workbook, worksheet
import converter as cnv

INPUT_SHEET = "Input"
REPORT_SHEET = "Report"
DILATION_SHEET = "Dilation"
EIEUR_SHEET = "Ei-Eur"
RFKN_SHEET = "Rf-K-n"
FLAC_SHEET = "FLAC1"
OUTPUT_SHEET = "Summary by Type"
SORT_COLUMN_IDX = cnv.col_name_to_idx("CG")
LAST_COLUMN_IDX = SORT_COLUMN_IDX

MAPPINGS = [
    # sheet name, src cell 1, [src cell 2,] dest col
    {"sheet": INPUT_SHEET, "in1": "D6", "out": "B", "offset": 0},
    {"sheet": INPUT_SHEET, "in1": "D8", "out": "D", "offset": 0},
    {"sheet": INPUT_SHEET, "in1": "I10", "out": "F", "offset": 0},
    {"sheet": INPUT_SHEET, "in1": "D9", "out": "G", "offset": 0},
    {"sheet": INPUT_SHEET, "in1": "I6", "out": "H", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "F21", "out": "L", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "G21", "out": "L", "offset": 1},
    {"sheet": REPORT_SHEET, "in1": "H21", "out": "L", "offset": 2},
    {"sheet": REPORT_SHEET, "in1": "F20", "out": "T", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "G20", "out": "T", "offset": 1},
    {"sheet": REPORT_SHEET, "in1": "H20", "out": "T", "offset": 2},
    {"sheet": REPORT_SHEET, "in1": "Q171", "out": "AG", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "Q171", "out": "AG", "offset": 1},
    {"sheet": REPORT_SHEET, "in1": "Q171", "out": "AG", "offset": 2},
    {"sheet": REPORT_SHEET, "in1": "Q170", "out": "AH", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "Q170", "out": "AH", "offset": 1},
    {"sheet": REPORT_SHEET, "in1": "Q170", "out": "AH", "offset": 2},
    {"sheet": REPORT_SHEET, "in1": "Q163", "in2": "Q194", "out": "AJ", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "Q163", "in2": "Q194", "out": "AJ", "offset": 1},
    {"sheet": REPORT_SHEET, "in1": "Q163", "in2": "Q194", "out": "AJ", "offset": 2},
    {"sheet": REPORT_SHEET, "in1": "O164", "in2": "O195", "out": "AI", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "O164", "in2": "O195", "out": "AI", "offset": 1},
    {"sheet": REPORT_SHEET, "in1": "O164", "in2": "O195", "out": "AI", "offset": 2},
    {"sheet": INPUT_SHEET, "in1": "D11", "out": "AN", "offset": 0},
    {"sheet": DILATION_SHEET, "in1": "V4", "out": "AQ", "offset": 0},
    {"sheet": DILATION_SHEET, "in1": "V30", "out": "AQ", "offset": 1},
    {"sheet": DILATION_SHEET, "in1": "V51", "out": "AQ", "offset": 2},
    {"sheet": DILATION_SHEET, "in1": "V11", "out": "AR", "offset": 0},
    {"sheet": DILATION_SHEET, "in1": "V37", "out": "AR", "offset": 1},
    {"sheet": DILATION_SHEET, "in1": "V58", "out": "AR", "offset": 2},
    {"sheet": REPORT_SHEET, "in1": "F45", "out": "AS", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "G45", "out": "AS", "offset": 1},
    {"sheet": REPORT_SHEET, "in1": "H45", "out": "AS", "offset": 2},
    {"sheet": REPORT_SHEET, "in1": "F39", "out": "BB", "offset": 0},
    {"sheet": REPORT_SHEET, "in1": "G39", "out": "BB", "offset": 1},
    {"sheet": REPORT_SHEET, "in1": "H39", "out": "BB", "offset": 2},
    {"sheet": EIEUR_SHEET, "in1": "Y2", "out": "BC", "offset": 0},
    {"sheet": EIEUR_SHEET, "in1": "Y24", "out": "BC", "offset": 1},
    {"sheet": EIEUR_SHEET, "in1": "Y49", "out": "BC", "offset": 2},
    {"sheet": EIEUR_SHEET, "in1": "Y19", "out": "BE", "offset": 0},
    {"sheet": EIEUR_SHEET, "in1": "Y42", "out": "BE", "offset": 1},
    {"sheet": EIEUR_SHEET, "in1": "Y67", "out": "BE", "offset": 2},
    {"sheet": EIEUR_SHEET, "in1": "Y9", "out": "BH", "offset": 0},
    {"sheet": EIEUR_SHEET, "in1": "Y31", "out": "BH", "offset": 1},
    {"sheet": EIEUR_SHEET, "in1": "Y57", "out": "BH", "offset": 2},
    {"sheet": RFKN_SHEET, "in1": "G56", "out": "BS", "offset": 0},
    {"sheet": RFKN_SHEET, "in1": "G57", "out": "BT", "offset": 0},
    {"sheet": RFKN_SHEET, "in1": "G75", "out": "BU", "offset": 0},
    {"sheet": RFKN_SHEET, "in1": "G76", "out": "BV", "offset": 0},
    {"sheet": FLAC_SHEET, "in1": "AD22", "out": "BW", "offset": 0},
    {"sheet": FLAC_SHEET, "in1": "AD23", "out": "BW", "offset": 1},
    {"sheet": FLAC_SHEET, "in1": "AD24", "out": "BW", "offset": 2},
    {"sheet": INPUT_SHEET, "in1": "D11", "out": "CG", "offset": 0},
    {"sheet": INPUT_SHEET, "in1": "D11", "out": "CG", "offset": 1},
    {"sheet": INPUT_SHEET, "in1": "D11", "out": "CG", "offset": 2},
]

MERGE_COLS = [
    "B", "D", "F", "G", "H", "AN", "BS", "BT", "BU", "BV"
]

OK = messagebox.OK
CANCEL = messagebox.CANCEL

def extract_file(source, destination):
    log.info("Extracting data from `%s` to `%s` ...", source, destination)
    wb_out = load_workbook(destination)
    wb_out.active = wb_out[OUTPUT_SHEET]

    extract_one(source, destination, wb_out)

    save_workbook(wb_out, destination)
    log.info("Done file extracting")


def extract_dir(source, destination):
    log.info("Extracting data from '%s' to '%s' ...", source, destination)
    wb_out = load_workbook(destination)
    wb_out.active = wb_out[OUTPUT_SHEET]
    start_row = wb_out.active.max_row + 1

    # Get list of *.xlsx files
    excel_files = list_excel_files(source)
    if len(excel_files) == 0:
        log.warning("No excel files found in '%s'", source)
        return

    for fn in excel_files:
        extract_one(fn, destination, wb_out)

    end_row = wb_out.active.max_row
    sort_rows(wb_out.active, start_row, end_row, 1, LAST_COLUMN_IDX)

    save_workbook(wb_out, destination)
    log.info("Done directory extracting")

# Try twice
def rename_orig(destination):
    try:
        os.rename(destination, new_file_name(destination))
        return True
    except PermissionError:
        resp = messagebox.askretrycancel(message=f"File '{destination}' is opened in another application.\n" \
                                    "Either close other and retry or cancel")
        if resp == CANCEL:
            return False
        try:
            os.rename(destination, new_file_name(destination))
            return True
        except PermissionError:
            return False


def save_workbook(wb, destination):
    if not rename_orig(destination):
        return

    try:
        wb.save(destination)
    except Exception as e: # pylint: disable=broad-except
        log.error("Failed to save file '%s': %s", destination, e)
        os.rename(new_file_name(destination), destination)


def list_excel_files(dir_path):
    excel_files = []
    for fn in os.listdir(dir_path):
        full_path = os.path.join(dir_path, fn)
        if re.match(r'.+\.xls[bmx]?$', fn) and os.path.isfile(full_path):
            excel_files.append(full_path)

    return excel_files


def extract_one(source, destination, wb_out):
    wb_in = load_workbook(source, data_only=True)

    last_row = wb_out.active.max_row
    log.info("Extracting data from '%s' to '%s' from row %d ...", source, destination, last_row)

    for mapping in MAPPINGS:
        try:
            copy_one_value(wb_in, wb_out, last_row, mapping)
        except Exception as e: # pylint: disable=broad-except
            log.error("Failed to extract data from '%s': %s", source, e)

    for col in MERGE_COLS:
        merge_cells(wb_out, col, last_row)


def new_file_name(fname):
    base, ext = os.path.splitext(fname)
    return base + ".org" + ext


# def print_mappings():
#     for mapping in MAPPINGS:
#         entry = mapping["sheet"] + " (" + mapping["in1"]
#         if "in2" in mapping:
#             entry += "+" + mapping["in2"]
#         entry += ") -> " + mapping["out"]
#         print(entry)


def copy_one_value(wb_in, wb_out, last_row, mapping):
    v = cnv.may_be_convert(wb_in[mapping["sheet"]][mapping["in1"]].value)

    if "in2" in mapping:
        v2 = cnv.may_be_convert(wb_in[mapping["sheet"]][mapping["in2"]].value)
        if not math.isnan(v) and not math.isnan(v2):
            v = (v + v2) / 2.0

    v_conv = cnv.convert_na(v)
    if v_conv == cnv.N_A:
        log.warning("Couldn't extract value from sheet '%s', cell '%s'",
                     mapping["sheet"], {mapping["in1"]})

    wb_out.active.cell(
        row = last_row + mapping["offset"] + 1,
        column = cnv.col_name_to_idx(mapping["out"]),
        value = v_conv
    )


def merge_cells(wb_out, col, last_row):
    col_idx = cnv.col_name_to_idx(col)
    wb_out.active.merge_cells(start_row=last_row + 1, end_row=last_row + 3, start_column=col_idx, end_column=col_idx)


# Based on https://stackoverflow.com/questions/44767554/sorting-with-openpyxl
# Copying format: https://stackoverflow.com/questions/45433425/how-to-move-cell-range-in-openpyxl-with-its-properties-hyperlinks-formatting-e
def sort_rows(ws: worksheet, row_start: int, row_end: int, col_start: int, col_end: int):
    """ Sorts range in the sheet
        row_start   First row to be sorted
        row_end     Last row to be sorted (inclusive)
        col_start   Start column to be sorted
        col_end     End column to be sorted (inclusive)
    """

    col_name_start = cnv.col_idx_to_name(col_start)
    col_name_end = cnv.col_idx_to_name(col_end)

    org_range = col_name_start + str(row_start) + ':' + col_name_end + str(row_end)
    shift = ws.max_row + 1 - row_start

    # Move whole range to the bottom
    ws.move_range(org_range, rows = shift)

    keys=[]
    for i in range(row_start + shift, row_end + shift + 1):
        val = ws.cell(row=i, column=SORT_COLUMN_IDX).value or ""
        keys.append((val, i))

    # Sorts by all fields in the tuple!
    keys.sort()

    # Move rows from original place to destination in sorted order
    to_row = row_start
    for key in keys:
        row_range = col_name_start + str(key[1]) + ':' + col_name_end + str(key[1])
        ws.move_range(row_range, rows = to_row - key[1])
        to_row += 1
